import openpyxl
import os
os.chdir('T:\Section of Hospital Medicine\miscellaneous\Research\ECRIP\Franak\Chris')
# Load workbook
wb = openpyxl.load_workbook('INV+Amb Frequencies_fbtest.xlsx')

nspre = wb.get_sheet_by_name('INV NS Pre')
# Define lists of terms we're looking for
refused = ['refused', 'ref', 'REFUSED']
bedrest = ['T&P', 'positioned', 'bed', 'bedrest', 'rest', 'in bed', 'awaiting orders']
totalassist = ['four', '4', 'hoyer', 'lift', 'full assist', 'total assist', '3 person', '3-person', 'three-person', 'three person']
twoperson = ['2 person', '2-person', 'two person', 'two-person', '2person', 'twoperson']
oneperson = ['1 person', '1-person', 'one person', 'one-person', '1person', 'oneperson']
standbyassist = ['stand by assist']
independent = ['wbat', 'gait belt', 'cane', 'walker', 'oob', 'up in room', 'up ad lib', 'ambulated', 'amb', 'ambulated', 'chair', 'bathroom', 'commode', 'independent', 'independently']
devices = ['cane', 'walker']
distanceambulated = ['ambulated', 'up in room', 'up ad lib']
distancebathroom = ['bathroom', 'commode']
distancechair = ['chair']
situp = ['sit up', 'dangle', 'sit up in bed']

print nspre.max_row

# Here's the actual algorithm that scans.
for i in range(2, nspre.max_row + 1):
    if not nspre["E" + str(i)].value:
        if any(x in nspre["A" + str(i)].value for x in refused):
            nspre["E" + str(i)].value = 999
        elif any(x in nspre["A" + str(i)].value for x in bedrest):
            if any(x in nspre["A" + str(i)].value for x in bedrest) and any(y in nspre["A" + str(i)].value for y in totalassist):
                nspre["E" + str(i)].value = 9
            elif any(x in nspre["A" + str(i)].value for x in bedrest) and any(y in nspre["A" + str(i)].value for y in twoperson):
                nspre["E" + str(i)].value = 8
            elif any(x in nspre["A" + str(i)].value for x in bedrest) and any(y in nspre["A" + str(i)].value for y in oneperson):
                nspre["E" + str(i)].value = 7
            elif any(x in nspre["A" + str(i)].value for x in situp):
                nspre["E" + str(i)].value = 6
            else:
                nspre["E" + str(i)].value = 10
        elif any(x in nspre["A" + str(i)].value for x in totalassist):
            nspre["E" + str(i)].value = 5
        elif any(x in nspre["A" + str(i)].value for x in twoperson):
            nspre["E" + str(i)].value = 4
        elif any(x in nspre["A" + str(i)].value for x in oneperson):
            nspre["E" + str(i)].value = 3
        elif any(x in nspre["A" + str(i)].value for x in standbyassist):
            nspre["E" + str(i)].value = 2
        elif any(x in nspre["A" + str(i)].value for x in independent):
            nspre["E" + str(i)].value = 1
    if nspre["E" + str(i)].value:
        if any(x in nspre["A" + str(i)].value for x in distanceambulated):
            nspre["F" + str(i)].value = "A"
        elif any(x in nspre["A" + str(i)].value for x in distancebathroom):
            nspre["F" + str(i)].value = "B"
        elif any(x in nspre["A" + str(i)].value for x in distancechair):
            nspre["F" + str(i)].value = "C"
    if nspre["E" + str(i)].value:
        if any(x in nspre["A" + str(i)].value for x in devices):
            nspre["G" + str(i)].value = "Y"
        else:
            nspre["G" + str(i)].value = 'N'

wb.save('test.xlsx')
